import os
import django
#import requests
import MySQLdb
from bs4 import BeautifulSoup
import openpyxl as xl
import click

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "onlineproject.settings")
django.setup()

from onlineapp.models import *


#####################################################################################
def extractWebData(soup):
    tableHeads = [x.text.strip("</th>").strip("<th>") for x in soup.find_all("th")][1:]
    tableData = [x.text.strip("</td>").strip("<td>").lstrip('0123456789') for x in soup.find_all("tr")][1:]


    return (tableHeads,tableData)

def storeWebDataToExcel(destName, tableHeads, tableData):
    workbook = xl.Workbook()
    workbook.create_sheet("TableData")
    workbook.save(destName)

    sheet = workbook['TableData']

    for col in range(6):
        cell = sheet.cell(row=1, column=col + 1)
        cell.value = tableHeads[col]


    for row in range(1,len(tableData)):
        inputs = tableData[row].split()
        for col in range(len(inputs)):
            cell = sheet.cell(row = row+1,column = col+1)
            cell.value = inputs[col]


    workbook.save(destName)

def read_from_xl(workbookName, sheetName):
    '''this function reads data from a specified sheet and returns a 2d list of data'''
    wb = xl.load_workbook(workbookName)
    sheetsrc = wb[sheetName]

    rows = sheetsrc.max_row
    cols = sheetsrc.max_column

    data = []
    for r in range(0, rows):
        data.append([])

    for r in range(rows):
        for c in range(cols):
            cell = sheetsrc.cell(row=r + 1, column=c + 1)
            data[r].append(cell.value)

    return data
#####################################################################################

# print(read_from_xl("students.xlsx","Colleges"))

@click.group()
def cli():
    pass


@cli.command('createdb')
def createDB():
    # a function to create a database and add 2 tables into it
    # db = MySQLdb.connect(host="localhost", user="shivasairam", passwd="shivasairam")
    # c = db.cursor()
    # c.execute("create database onlineappschema")
    # db.commit()
    print('created db')


@cli.command('deletedb')
def deleteDB():
    db = MySQLdb.connect(host="localhost", user="shivasairam", passwd="shivasairam")
    c = db.cursor()
    c.execute("drop database onlineappschema")
    db.commit()
    print('deleted db')


@cli.command('populatedb')
@click.argument('stdFile')
@click.argument('mrksFile')
def populatedb(stdfile,mrksfile):
    # request = requests.get(mrksfile).text
    # soup = BeautifulSoup(request, 'html5lib')
    #
    # webData = extractWebData(soup)
    # storeWebDataToExcel("result1.xlsx", webData[0], webData[1])

    marks = read_from_xl("result1.xlsx", "TableData")[1:]
    colleges = read_from_xl(stdfile, "Colleges")[1:]
    current = read_from_xl(stdfile, "Current")[1:]
    deletions = read_from_xl(stdfile, "Deletions")[1:]

    collegeDict = dict()
    stdFolderDict = dict()

    for college in colleges:
        c = College(name=college[0], location=college[2], acronym=college[1], contact=college[3])
        c.save()
        collegeDict[college[1]] = c

    for student in current:
        try:
            collegeDict[student[1]]
        except Exception:
            print(student[1], "not found")
        else:
            s = Student(name=student[0], dob=None, email=student[2], db_folder=student[3], dropped_out=False,
                        college=collegeDict[student[1]])
            s.save()
            stdFolderDict[str(student[3]).lower()] = s

    for dStudent in deletions:
        s = Student(name=dStudent[0], dob=None, email=dStudent[2], db_folder=dStudent[3], dropped_out=True,
                    college=collegeDict[dStudent[1]])
        s.save()


    for std in marks:
        dbname = str(str(std[0]).split('_')[2]).lower()
        try:
            stdFolderDict[dbname]
        except Exception:
            print(dbname, " not found")
        else:
            m = MockTest1(problem1=str(std[1]), problem2=str(std[2]), problem3=str(std[3]), problem4=str(std[4]),
                          total=str(std[5]), students=stdFolderDict[dbname])
            m.save()


@cli.command('cleardata')
def clearData():
    College.delete()
    print('cleardata')


if __name__ == '__main__':
    cli()


# c = College(name = "Vasavi college",location = "Hyderabad",acronym = "vce",contact = "vce@gmail.com")
# c.save()

