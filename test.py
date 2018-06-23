import os

import MySQLdb
import click
import django
from django.core.management import call_command
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "onlineproject.settings")
django.setup()
from onlineapp.models import *

username = "shivasairam"
password = "shivasairam"
db_name = "onlineapp"
server = "localhost"
# prepare a cursor object using cursor() method
db = MySQLdb.connect(server, username, password, db_name)
cursor = db.cursor()


@click.group()
def cli():
    pass


def get_college(name, wb):
    sheet = wb["Colleges"]
    i = 1
    for row in sheet:
        if sheet.cell(row=i, column=2).value == name:
            try:
                c = College(name=sheet.cell(row=i, column=1).value, acronym=sheet.cell(row=i, column=2).value,
                            location=sheet.cell(row=i, column=3).value, contact=sheet.cell(row=i, column=4).value)
                c.save()
                return c
            except ValueError:
                pass
            except django.db.utils.IntegrityError:
                #  print(name)
                return College.objects.get(acronym=name)
        i += 1
    return None


def process_students(_sheet, wb):
    size = 0
    for i in _sheet:
        size += 1
    with click.progressbar(_sheet,
                           label='Populating Students',
                           length=size) as sheet:
        for row in sheet:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            clg = get_college(tmp[1], wb)
            if clg is None:
                continue
            s = Student(name=tmp[0], college=clg, email=tmp[2], db_folder=tmp[3])
            s.save()
        sheet = wb["Deletions"]
        for row in sheet:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            clg = get_college(tmp[1], wb)
            if clg is None:
                continue
            s = Student(name=tmp[0], dropped_out=1, college=clg, email=tmp[2], db_folder=tmp[3])
            s.save()


def get_student(online_name):
    dbname = online_name[online_name.find("_", online_name.find("_") + 1) + 1:-5]
    try:
        s = Student.objects.get(db_folder=dbname)
        return s
    except Exception:
        pass


def process_marks(wb):
    _sheet = wb.active
    size = 0
    for i in _sheet:
        size += 1
    with click.progressbar(_sheet,
                           label='Populating Mock Test marks',
                           length=size) as sheet:
        for row in sheet:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            s = get_student(tmp[0])
            if s is None:
                continue
            m = Mocktest(problem1=tmp[1],
                         problem2=tmp[2],
                         problem3=tmp[3],
                         problem4=tmp[4],
                         totals= int(tmp[1]) + int(tmp[2]) + int(tmp[3]) + int(tmp[4]),
                         student=s)
            m.save()



@cli.command()
@click.argument("students_xl", nargs=1)
@click.argument("marks_xl", nargs=1)
def populatedb(students_xl, marks_xl):
    """Populates data from excel sheet to database"""
    wb = load_workbook(students_xl)
    process_students(wb["Current"], wb)
    wb = load_workbook(marks_xl)
    process_marks(wb)


@cli.command()
def createdb():
    call_command('migrate')


@cli.command()
def cleardata():
    """Deletes all rows"""
    College.objects.all().delete()
    Student.objects.all().delete()
    Mocktest.objects.all().delete()


@cli.command()
def dropdb():
    """Deletes all tables"""
    call_command('migrate onlineapp zero')


if __name__ == "__main__":
    # createdb()
    cli()
