import os

import MySQLdb
import click
import django
from openpyxl import load_workbook

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "onlineproject.settings")
django.setup()
from onlineapp.models import *

username = "root"
password = "1234"
db_name = "mydb"
server = "localhost"
# prepare a cursor object using cursor() method
db = MySQLdb.connect(server, username, password, db_name)
cursor = db.cursor()

wb = load_workbook("students.xlsx")


@click.group()
def cli():
    pass


def get_college(name):
    sheet = wb["Colleges"]
    i = 1
    for row in sheet:
        if sheet.cell(row=i, column=2).value == name:
            try:
                c = College(name=sheet.cell(row=i, column=1).value, acronym=sheet.cell(row=i, column=2).value,
                            location=sheet.cell(row=i, column=3).value, contact=sheet.cell(row=i, column=4).value)
                c.save()
                return c
            except ValueError:
                pass
            except django.db.utils.IntegrityError:
                #  print(name)
                return College.objects.get(acronym=name)
        i += 1
    return None


def process_students(_sheet):
    size = 0;
    for i in _sheet:
        size += 1
    with click.progressbar(_sheet,
                           label='Populating Students',
                           length=size) as sheet:
        for row in sheet:
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            clg = get_college(tmp[1])
            if clg is None:
                continue
            s = Student(name=tmp[0], college=clg, email=tmp[2], db_folder=tmp[3])
            s.save()


@cli.command()
@click.argument("students_xl", default="students.xlsx", nargs=1)
@click.argument("marks_xl", default="marks.xlsx", nargs=1)
def populatedb(students_xl, marks_xl):
    """Populates data from excel sheet to database"""
    process_students(wb["Current"])


@cli.command()
def createdb():
    """Creates new database"""
    # cursor.execute("CREATE DATABASE mydb")

    cursor.execute("""CREATE TABLE `onlineapp_college` (
    `id` integer AUTO_INCREMENT NOT NULL PRIMARY KEY,
    `name` varchar(128) NOT NULL,
    `location` varchar(64) NOT NULL,
    `acronym` varchar (8) NOT NULL UNIQUE,
    `contact` varchar(254) NOT NULL);""")

    cursor.execute("""CREATE TABLE `onlineapp_student` (`id` integer AUTO_INCREMENT NOT NULL PRIMARY KEY, 
    `name` varchar(128) NOT NULL, `dob` date NULL, 
    `email` varchar(254) NOT NULL, 
    `db_folder` varchar(50) NOT NULL, 
    `dropped_out` bool NOT NULL, 
    `collge_id` integer NOT NULL  FOREIGN KEY(`collge_id`) REFERENCES `onlineapp_college` (`id`);;""")
    cursor.execute("""CREATE TABLE `onlineapp_mocktest` (`id` integer AUTO_INCREMENT NOT NULL PRIMARY KEY, `problem1` integer NOT NULL, `problem2` integer NOT NULL, `problem3` integer NO
    T NULL, `problem4` integer NOT NULL, `totals` integer NOT NULL,student_id integer NOT NULL FOREIGN KEY FOREIGN KEY (`student_id`) REFERENCES `onlineapp_st
    udent` (`id`);""")
    db.commit()


@cli.command()
def cleardata():
    """Deletes all rows"""
    dropdb()
    createdb()


@cli.command()
def dropdb():
    """Deletes all tables"""
    cursor.execute("DROP TABLE IF EXISTS online_students")
    cursor.execute("DROP TABLE IF EXISTS online_mocktest")
    cursor.execute("DROP TABLE IF EXISTS online_college")
    # cursor.execute("drop database mydb")


if __name__ == "__main__":
    # createdb()
    cli()
